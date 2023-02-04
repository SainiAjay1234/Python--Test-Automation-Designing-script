import pywinauto
import time
from pywinauto import application
import pyautogui
import subprocess
import pyautogui
import os
from os import sys
import pandas as pd
import xlrd
import xlrd3
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from subprocess import Popen




wrkbk = load_workbook("C:\\Automation\\TestCaseInputFile2.xlsx")
# to identify the active sheet
sh = wrkbk.active
# get the value of row 2 and column 3
c1=sh.cell(row=2,column=7)
c2=sh.cell(row=2,column=8)
c3=sh.cell(row=2,column=9)
c4=sh.cell(row=2,column=10)
c5=sh.cell(row=2,column=11)
c6=sh.cell(row=2,column=12)
c7=sh.cell(row=2,column=13)




ConfigMgrIPAddress = c1.value
OVD = c2.value
NWDR = c3.value
DEV = c4.value
MOD3 = c5.value
NWPNT = c6.value
OUT = c7.value



app = pywinauto.application.Application(backend = "uia")
app1 = pywinauto.application.Application(backend = "uia")


class VRNISecond:
    def LaunchVRNI(self,app):     # To Launch VRNI Configuration manager
          app = app.start("C:\Ovation\OvationBase\OvVirtualRNIConfigMgr.exe")
          app.OvationVirtualRNIConfigurationManager.wait('ready', timeout = 3000)
          app.OvationVirtualRNIConfigurationManager.set_focus()
          time.sleep(3)
    def ConnectVRNI(self,app):    # To Connect VRNI by entering VRNI host IP Address
         from pywinauto import application
         dlg2 = app.OvationVirtualRNIConfigurationManager.child_window(title=ConfigMgrIPAddress, control_type="Edit")
         time.sleep(2)
         dlg2.type_keys(ConfigMgrIPAddress+"{DEL 15}")
         dlg2.type_keys("{TAB}{ENTER}")
         time.sleep(2)
    def AddVRNIDrop(self,app):  # To add a New Drop to VRNI Configuration manager
         dlg1 = app.OvationVirtualRNIConfigurationManager
         dlg4 = dlg1.child_window(title=OVD, control_type="TreeItem")
         dlg4.select()
         dlg4.type_keys("{TAB}{TAB}{SPACE}")
         dlg5 = dlg1.child_window(title=NWDR, control_type="Edit")
         dlg5.select()
         dlg5.select().click_input(button='left')
         dlg5.type_keys("{BACKSPACE 8}"+"Drop6"+"{TAB}{DELETE}"+"6"+"{TAB}"+"192.168.104.6"+"{TAB}{TAB}{TAB}{ENTER}"+"{TAB}{ENTER}")
         Test = app.OvationVirtualRNIConfigurationManager.Drop6.select()
         time.sleep(2)
    def  AnalogValueInject(self,app):   # To inject Analog point value which is residing inside  3rd module 
          Test = app.OvationVirtualRNIConfigurationManager.Drop6.select()
          Test.type_keys('{RIGHT}'+'{DOWN}')
          dlg1 = app.OvationVirtualRNIConfigurationManager
          dlg7 = dlg1.child_window(title=DEV, control_type="TreeItem")
          dlg7.Device1.double_click_input(button='left')
          time.sleep(3)
          dlg7.type_keys('{DOWN}'+'+{RIGHT}'+'{TAB}{TAB}{TAB}{TAB}{SPACE}')
          dlg9 = dlg1.child_window(title=MOD3, control_type="Text")
          dlg9.set_focus().click_input(button='left', double='true',coords=(0, 0))
          time.sleep(2)
          dlg9.type_keys('{TAB 5}'+'{ENTER}'+'{TAB 6}{DOWN}'+'{RIGHT}'+'{DOWN}')
          time.sleep(3)
          dlg10 = dlg1.child_window(title="0", control_type="Edit").wrapper_object()
          dlg10.select()
          dlg10.type_keys("50.01")
          time.sleep(3)
          dlg10.type_keys('{TAB}{ENTER}')
          time.sleep(2)
    def  ReadAnalogValue(self,app1):   # To read  Analog point value from PI and compare it with injected analog point value and printed result in a file
          app1 = app.start("C:\Ovation\OvationBase\PI.exe")
          app1.OvationPointInformation.wait('ready', timeout = 30)
          app1.OvationPointInformation.set_focus()
          time.sleep(5)
          dlg = app1.OvationPointInformation
          dlg1 = dlg.child_window(auto_id=NWPNT, control_type="Pane")
          dlg1.click_input()
          dlg1.type_keys("ANALOG_2"+"~")
          var1 = dlg.child_window(title="50.01", control_type="Custom").exists(1)
          if var1 == True :
            with open(OUT,'w')as f : print ("ANALOG_A2 value changed to 50",file=f)
          else:
           with open(OUT,'w')as f :  print ("Test case is failed to change the value",file=f)

         
         
          
h2 = VRNISecond()

h2.LaunchVRNI(app)
time.sleep(3)
h2.ConnectVRNI(app)
time.sleep(3)
h2.AddVRNIDrop(app)
time.sleep(3)
h2.AnalogValueInject(app)
time.sleep(3)
h2.ReadAnalogValue(app1)
