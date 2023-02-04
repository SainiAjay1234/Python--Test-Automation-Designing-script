import pywinauto
import time
from pywinauto import application
import pyautogui
import subprocess
import pyautogui
import os
from os import sys
import pandas as pd
import xlrd
import xlrd3
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from subprocess import Popen



wrkbk = load_workbook("C:\\Automation\\TestCaseInputFile.xlsx")
# to identify the active sheet
sh = wrkbk.active
# get the value of row 2 and column 7
c1=sh.cell(row=2,column=7)
c2=sh.cell(row=2,column=8)
c3=sh.cell(row=2,column=9)
c4=sh.cell(row=2,column=10)
c5=sh.cell(row=2,column=11)
c6=sh.cell(row=2,column=12)
c7=sh.cell(row=2,column=13)
c8=sh.cell(row=2,column=14)
c9=sh.cell(row=2,column=15)


ConfigMgrIPAddress = c1.value
OVD =  c2.value
NWDR = c3.value
DEV =  c4.value
MOD1 =  c5.value
PIINJ = c6.value
NWPNT = c7.value
PARA = c8.value
OUT = c9.value


app = pywinauto.application.Application(backend = "uia")
app1 = pywinauto.application.Application(backend = "uia")


class VRNIFirst:
    def LaunchVRNI(self,app):   # To Launch VRNI Configuration manager
          app = app.start("C:\Ovation\OvationBase\OvVirtualRNIConfigMgr.exe")
          app.OvationVirtualRNIConfigurationManager.wait('ready', timeout = 3000)
          app.OvationVirtualRNIConfigurationManager.set_focus()
          time.sleep(3)
    def ConnectVRNI(self,app):   # To Connect VRNI by entering VRNI host IP Address
         dlg2 = app.OvationVirtualRNIConfigurationManager.child_window(title=ConfigMgrIPAddress, control_type="Edit")
         time.sleep(2)
         dlg2.type_keys(ConfigMgrIPAddress+"{DEL 15}")
         dlg2.type_keys("{TAB}{ENTER}")
         time.sleep(2)
    def AddVRNIDrop(self,app):   # To add a New Drop to VRNI Configuration manager
         dlg1 = app.OvationVirtualRNIConfigurationManager
         dlg4 = dlg1.child_window(title=OVD, control_type="TreeItem")
         dlg4.select()
         dlg4.type_keys("{TAB}{TAB}{SPACE}")
         dlg5 = dlg1.child_window(title=NWDR, control_type="Edit")
         dlg5.select()
         dlg5.select().click_input(button='left')
         dlg5.type_keys("{BACKSPACE 8}"+"Drop6"+"{TAB}{DELETE}"+"6"+"{TAB}"+"192.168.104.6"+"{TAB}{TAB}{TAB}{ENTER}"+"{TAB}{ENTER}")
         Test = app.OvationVirtualRNIConfigurationManager.Drop6.select()
         time.sleep(2)
    def  DigitalValueInject(self,app):  # To inject Digital point value which is residing inside module 1st
          Test = app.OvationVirtualRNIConfigurationManager.Drop6.select()
          Test.type_keys('{RIGHT}'+'{DOWN}')
          dlg1 = app.OvationVirtualRNIConfigurationManager
          dlg7 = dlg1.child_window(title=DEV, control_type="TreeItem")
          dlg7.Device1.double_click_input(button='left')
          dlg7.type_keys('{DOWN}'+'+{RIGHT}'+'{TAB}{TAB}{TAB}{TAB}{SPACE}')
          time.sleep(2)
          dlg8 = dlg1.child_window(title=MOD1, control_type="Text")
          dlg8.set_focus().click_input(button='left', double='true',coords=(0, 0))
          time.sleep(2)
          dlg8.type_keys('{TAB 5}'+'{ENTER}'+'{TAB 6}{DOWN}'+'{RIGHT}'+'{DOWN}') 
          time.sleep(3)
          dlg9 = dlg1.child_window(auto_id="DigitalValueCombo", control_type="ComboBox")
          dlg9.double_click_input(button='left')
          time.sleep(2)
          dlg9.type_keys('{DOWN}')
          time.sleep(2)
          dlg10 = dlg1.child_window(title="Inject", control_type="Button").wrapper_object()
          dlg10.click_input()
    def  ReadDigitalValue(self,app1):  # To read  Digital point value from PI and compare it with injected digital point value and printed result in a file
          app1 = app.start("C:\Ovation\OvationBase\PI.exe")
          app1.OvationPointInformation.wait('ready', timeout = 30)
          app1.OvationPointInformation.set_focus()
          time.sleep(5)
          dlg = app1.OvationPointInformation
          dlg1 = dlg.child_window(auto_id=NWPNT, control_type="Pane")
          dlg1.click_input()
          dlg1.type_keys("DIGITAL_D1"+"~")
          dlg = app1.OvationPointInformation
          var1 = dlg.child_window(title=PARA, control_type="DataItem").exists(1)
          if var1 == True :
            with open(OUT,'w')as f : print ("Testcase passed digital value changed to 1" ,file=f)
          else:
           with open(OUT,'w')as f :  print ("Test case is failed to change the value",file=f)

         
         


h2 = VRNIFirst()

h2.LaunchVRNI(app)
time.sleep(3)

h2.ConnectVRNI(app)
time.sleep(3)

h2.AddVRNIDrop(app)
time.sleep(3)

h2.DigitalValueInject(app)
time.sleep(3)

h2.ReadDigitalValue(app1)
